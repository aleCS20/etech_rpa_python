# src/transformers/template_injector.py
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from pathlib import Path

class TemplateInjector:
    def __init__(self, pasta_template: Path, pasta_destino: Path):
        self.pasta_template = pasta_template
        self.pasta_destino = pasta_destino

    def injetar_dados_no_gabarito_xlsx(self, nome_template, nome_saida_xlsx, nome_guia, dataframe_dados):
        """Injeta os dados no template .xlsx mantendo as cores e estilos impecáveis."""
        caminho_template = self.pasta_template / nome_template
        caminho_saida_xlsx = self.pasta_destino / nome_saida_xlsx

        print(f"📋 [Injector] Carregando a cópia estável do Gabarito: {nome_template}")
        wb = openpyxl.load_workbook(str(caminho_template), data_only=False)
        
        if nome_guia in wb.sheetnames:
            aba = wb[nome_guia]
            print(f"   ↳ Guia '{nome_guia}' selecionada via OpenPyXL.")
        else:
            aba = wb.active

        # Limpa dados antigos da segunda linha para baixo (preservando o cabeçalho e estilos originais)
        if aba.max_row > 1:
            print("🧹 [Injector] Limpando dados antigos mantendo as linhas de estilo ativas...")
            aba.delete_rows(2, aba.max_row)

        print("⚡ [Injector] Despejando registros filtrados na planilha...")
        # Transforma o DataFrame em linhas e adiciona na aba mantendo a formatação padrão da tabela
        for row in dataframe_to_rows(dataframe_dados, index=False, header=False):
            aba.append(row)

        # Salva o arquivo temporário em formato .xlsx
        wb.save(str(caminho_saida_xlsx))
        wb.close()
        print(f"✅ [Injector] Arquivo intermediário salvo com sucesso em: {nome_saida_xlsx}")
        return caminho_saida_xlsx
