import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

class ExcelInjector:
    """Classe especialista em injetar matrizes de dados dentro de templates OpenPyXL."""
    def __init__(self, pasta_template: Path, pasta_processed: Path):
        self.pasta_template = pasta_template
        self.pasta_processed = pasta_processed

    def injetar_no_template_xlsx(self, nome_template, nome_saida, nome_guia, dataframe_dados):
        caminho_template = self.pasta_template / nome_template
        caminho_final = self.pasta_processed / nome_saida

        print(f"📋 [Injector] Abrindo o gabarito oficial da empresa: {nome_template}")
        wb = openpyxl.load_workbook(str(caminho_template), data_only=False)
        
        if nome_guia in wb.sheetnames:
            aba = wb[nome_guia]
            print(f"   ↳ Guia selecionada: '{nome_guia}'")
        else:
            aba = wb.active
            print(f"⚠️ Aba '{nome_guia}' não localizada no teste. Injetando na aba padrão ativa.")

        # =====================================================================
        # 🛠️ AJUSTE DEFINITIVO: PROTEÇÃO DE CABEÇALHOS DE DUAS LINHAS
        # =====================================================================
        # Limpa apenas o TEXTO das linhas de dados antigos a partir da LINHA 3.
        # Linha 1 (Título) e Linha 2 (Títulos de Colunas) permanecem intactos.
        # # Como o cabeçalho original ocupa apenas a Linha 1, passamos apagando
        # os dados antigos restritamente da linha 2 para baixo, mantendo os títulos ajuste para a indústria. > 1
        max_linhas_gabarito = aba.max_row
        if max_linhas_gabarito > 2:
            print("🧹 [Injector] Apagando registros antigos a partir da Linha 3...")
            for row_idx in range(3, max_linhas_gabarito + 1):
                for col_idx in range(1, aba.max_column + 1):
                    aba.cell(row=row_idx, column=col_idx).value = None

        print("⚡ [Injector] Injetando novos dados a partir da Linha 3 com formatação original...")
        
        # Converte o DataFrame limpo do Pandas em uma matriz pura
        registros_novos = dataframe_dados.values.tolist()

        # =====================================================================
        # 🛠️ AJUSTE PARA A INDÚSTRIA: ESCRITA A PARTIR DA LINHA 2
        # =====================================================================
        # Configura o início da gravação dos dados reais estritamente para a Linha 3
        linha_atual = 3
        for linha_dados in registros_novos:
            for col_idx, valor in enumerate(linha_dados, start=1):
                celula = aba.cell(row=linha_atual, column=col_idx)
                
                # Modifica apenas o valor da célula. O design (fontes/grades) original é mantido.
                celula.value = "" if pd.isna(valor) else str(valor).strip()
            
            linha_atual += 1

        # Salva o arquivo final consolidado pronto para produção
        print(f"💾 [Injector] Gravando planilha final atualizada: /processed/{nome_saida}")
        wb.save(str(caminho_final))
        wb.close()
        print("✅ [Injector] Injeção executada mantendo os cabeçalhos e estilos!")


